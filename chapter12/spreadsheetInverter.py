#! /usr/bin/python3
# spreadsheetInverter.py - Inverts the row and column of the cells in the spreadsheet.

import openpyxl, sys
from openpyxl.utils import get_column_letter

data = []
wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.active
newWb = openpyxl.Workbook()
newSheet = newWb.active

data = tuple(sheet.rows)

for r in range(1, sheet.max_column + 1):
    for c in range(1, sheet.max_row + 1):
        newSheet.cell(row=r, column=c).value = data[c - 1][r - 1].value

newWb.save('inverted_' + sys.argv[1])