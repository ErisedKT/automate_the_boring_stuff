#! /usr/bin/python3
# multiplicationTable.py - Takes a number N and creates an NxN 
# muliplication table in an Excel spreadsheet.

# Usage: ./multiplicationTable.py <N>

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Check if N is a valid argument.

N = int(sys.argv[1])
if N < 1:
    print('Incorrect argument!')
    sys.exit()

# Create an empty spreadsheet.
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

# Create Font object for labels.
boldText = Font(bold=True, name='Calibri')
# Freeze label row and columns.
sheet.freeze_panes = 'B2'

# Create label row and column.
for i in range(1, N + 1):
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=1, column=i+1).value = i

# Apply Font object to labels.
for cell in sheet['1:1']:
    cell.font = boldText

for cell in sheet['A:A']:
    cell.font = boldText

# Fill in multiplied products to cells.
for rowOfCellObj in sheet['B2:' + get_column_letter(N + 1) + str(N + 1)]:
    for cellObj in rowOfCellObj:
        cellObj.value = (cellObj.row - 1) * (cellObj.column - 1)

# Save spreadsheet.
wb.save('times%s.xlsx' % N)